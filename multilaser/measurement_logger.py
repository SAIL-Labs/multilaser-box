"""
Measurement Logger
Logs power meter measurements to Excel (.xlsx) or CSV (.csv) files.

Two Excel layouts are supported, both generated programmatically:

- "throughput" (default): the SAIL lantern-fabrication throughput layout.
  Columns A-D hold Trial, Port, injection power and lantern output power,
  while prefilled formulas in columns E-G and I-T compute throughput, loss
  and per-port statistics. Rows are appended per trial.

- "report": the SAIL Lantern Test Report layout. One sheet per wavelength
  (1550/1310/1064 nm), each with a header block (Lantern S/N, sources,
  PMREF in B17 [uW], Launch power in B18 [mW]) and a port-keyed table
  (rows 22-40 = ports 1-19) holding Throughput (mW) in C and the raw
  reference reading (uW) in D; formulas compute calibrated insertion loss
  and throughput as C/(B18*D/B17). Logging a port again overwrites its row.

CSV logs store the throughput-layout data with throughput and loss computed
at log time.

Requirements:
- openpyxl (optional): pip install openpyxl  -- required for Excel logging only

Author: Multi-Laser Box Project
Date: 2026-07-13
"""

import csv
import logging
import math
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.worksheet.formula import ArrayFormula

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

CSV_HEADER = [
    "timestamp",
    "trial",
    "port",
    "injection_power_W",
    "lantern_output_power_W",
    "throughput",
    "loss_percent",
    "loss_dB",
    "wavelength_nm",
]

# Throughput worksheet layout constants (1-indexed openpyxl columns)
COL_TRIAL = 1  # A
COL_PORT = 2  # B
COL_INJECTION = 3  # C
COL_OUTPUT = 4  # D
COL_THROUGHPUT = 5  # E
FIRST_DATA_ROW = 2
PREFILLED_ROWS = 415  # rows with throughput/loss formulas prefilled
SUMMARY_PORTS = 7  # per-port statistics rows (I2:O8)

XLSX_HEADERS = {
    "A1": "Trial",
    "B1": "Port",
    "C1": "injection power",
    "D1": "lantern output power",
    "E1": "throughput",
    "F1": "loss %",
    "G1": "loss dB",
    "I1": "Port",
    "J1": "injection power",
    "K1": "lantern output power",
    "L1": "throughput",
    "M1": "throughput std",
    "N1": "loss %",
    "O1": "loss dB",
    "Q1": "Overall Throughput",
    "R1": "Std",
    "S1": "Overall Loss",
    "T1": "Overall Average dB",
}

XLSX_COLUMN_WIDTHS = {
    "C": 24.5,
    "D": 21.2,
    "E": 12.5,
    "F": 11.0,
    "G": 11.0,
    "P": 10.2,
    "Q": 12.0,
}

# Lantern Test Report layout constants
REPORT_WAVELENGTHS_NM = [1550, 1310, 1064]  # one sheet per wavelength
REPORT_PMREF_CELL = "B17"  # reference reading at launch measurement (uW)
REPORT_LAUNCH_CELL = "B18"  # launch power via ref patch cord into PM1 (mW)
REPORT_HEADER_ROW = 21
REPORT_FIRST_PORT_ROW = 22
REPORT_MAX_PORTS = 19  # table rows 22-40
REPORT_COL_PORT = 1  # A
REPORT_COL_CONNECTOR = 2  # B
REPORT_COL_OUTPUT_MW = 3  # C: lantern output "Throughput (mW)"
REPORT_COL_REF_UW = 4  # D: raw reference reading (uW)

REPORT_COLUMN_WIDTHS = {
    "A": 12.9,
    "B": 13.7,
    "C": 17.0,
    "D": 14.1,
    "E": 27.6,
    "F": 23.6,
    "G": 19.9,
    "H": 16.9,
    "I": 28.3,
    "J": 23.6,
}


class MeasurementLogError(Exception):
    """Exception raised for measurement logging errors"""

    pass


class MeasurementLogger:
    """Appends power measurements to an Excel or CSV log file.

    The file format is inferred from the file extension (.xlsx or .csv).
    The file is opened, written and closed on every append so it is safe
    to inspect between measurements (but must not be open in Excel while
    logging).
    """

    def __init__(self, file_path: str, layout: Optional[str] = None):
        self.file_path = Path(file_path)
        suffix = self.file_path.suffix.lower()
        # 'throughput' or 'report' for xlsx; resolved lazily from the file
        # contents for existing files when not given
        self.layout = layout

        if suffix == ".xlsx":
            if not OPENPYXL_AVAILABLE:
                raise MeasurementLogError(
                    "Excel logging requires the openpyxl package.\n"
                    "Install it with: pip install openpyxl\n"
                    "Alternatively, use a .csv log file."
                )
            self.format = "xlsx"
        elif suffix == ".csv":
            self.format = "csv"
            self.layout = "throughput"
        else:
            raise MeasurementLogError(
                f"Unsupported log file type: '{suffix}' (use .xlsx or .csv)"
            )

    def create_new(
        self,
        wavelength_nm: Optional[int] = None,
        layout: str = "throughput",
        serial: Optional[str] = None,
    ):
        """Create a new log file, overwriting any existing file.

        Excel logs are generated with the throughput worksheet layout on a
        sheet named 'wave_<wavelength>', or with the Lantern Test Report
        layout (layout='report', one sheet per wavelength, optional lantern
        serial number). CSV logs are created with a header row.
        """
        if self.format == "xlsx":
            if layout == "report":
                self.layout = "report"
                self._create_report_xlsx(serial)
            else:
                self.layout = "throughput"
                self._create_xlsx(wavelength_nm)
        else:
            self._create_csv()
        logger.info(f"Created measurement log: {self.file_path}")

    def append_measurement(
        self,
        port: int,
        injection_power_w: Optional[float],
        output_power_w: Optional[float],
        wavelength_nm: Optional[int] = None,
        raw_reference_w: Optional[float] = None,
    ) -> int:
        """Record one measurement and return its trial (or port) number.

        Throughput/CSV logs append a new trial row using the corrected
        injection power. Report logs write output power (mW) and the raw
        reference reading (uW) into the port's row on the sheet nearest
        wavelength_nm, overwriting any previous values for that port.
        """
        if self.format == "xlsx":
            if self._resolve_layout() == "report":
                return self._append_report_xlsx(
                    port, output_power_w, raw_reference_w, wavelength_nm
                )
            return self._append_xlsx(port, injection_power_w, output_power_w)
        return self._append_csv(port, injection_power_w, output_power_w, wavelength_nm)

    def read_measurements(self, wavelength_nm: Optional[int] = None):
        """Return logged measurements as (trial, port, injection_w, output_w) tuples.

        Returns an empty list if the file does not exist yet. Rows with no
        data in the trial/port/power columns are skipped. For report logs,
        rows come from the sheet nearest wavelength_nm (first sheet if None),
        injection power is derived from the sheet's Launch/PMREF cells, and
        each tuple carries a 5th element: the raw reference reading (W)
        stored in the sheet.
        """
        if not self.file_path.exists():
            return []
        if self.format == "xlsx":
            if self._resolve_layout() == "report":
                return self._read_report_xlsx(wavelength_nm)
            return self._read_xlsx()
        return self._read_csv()

    def _resolve_layout(self) -> str:
        """Determine the xlsx layout, inspecting an existing file if needed"""
        if self.layout is None:
            if self.file_path.exists():
                wb = self._load_workbook(read_only=True)
                try:
                    ws = wb.worksheets[0]
                    is_report = (
                        ws["A1"].value == "Lantern S/N"
                        or ws[f"A{REPORT_HEADER_ROW}"].value == "Port Number"
                    )
                    self.layout = "report" if is_report else "throughput"
                finally:
                    wb.close()
            else:
                self.layout = "throughput"
        return self.layout

    def _load_workbook(self, read_only: bool = False, data_only: bool = False):
        """Open the workbook with the standard error handling"""
        try:
            return openpyxl.load_workbook(
                self.file_path, read_only=read_only, data_only=data_only
            )
        except PermissionError:
            raise MeasurementLogError(self._locked_file_message())
        except Exception as e:
            raise MeasurementLogError(f"Failed to open log file:\n{str(e)}")

    # === Excel implementation ===

    def _create_xlsx(self, wavelength_nm: Optional[int]):
        wb = openpyxl.Workbook()
        # openpyxl writes no cached formula results; make Excel recalculate
        wb.calculation.fullCalcOnLoad = True
        ws = wb.active
        ws.title = f"wave_{wavelength_nm}" if wavelength_nm else "throughput"

        header_font = Font(bold=True)
        for coord, header in XLSX_HEADERS.items():
            ws[coord] = header
            ws[coord].font = header_font
        for column, width in XLSX_COLUMN_WIDTHS.items():
            ws.column_dimensions[column].width = width

        # Per-measurement throughput/loss formulas (columns E-G)
        for row in range(FIRST_DATA_ROW, PREFILLED_ROWS + 1):
            self._write_row_formulas(ws, row)

        # Per-port statistics (columns I-O): the port list spills from a
        # UNIQUE/FILTER formula; _xlfn/_xlws prefixes are how openpyxl must
        # write post-2007 Excel function names
        last = FIRST_DATA_ROW + SUMMARY_PORTS - 1
        ws["I2"] = ArrayFormula(
            f"I2:I{last}",
            '=_xlfn.UNIQUE(_xlfn._xlws.FILTER(B:B, (B:B<>"") * (B:B<>"Port")))',
        )
        for row in range(FIRST_DATA_ROW, last + 1):
            for col, src in (("J", "C"), ("K", "D"), ("L", "E"), ("N", "F"), ("O", "G")):
                ws[f"{col}{row}"] = (
                    f'=IFERROR(AVERAGEIF($B:$B,$I{row},{src}:{src}),"")'
                )
            ws[f"M{row}"] = ArrayFormula(
                f"M{row}",
                f'=IFERROR(_xlfn.STDEV.S(_xlfn._xlws.FILTER(E:E,B:B=$I{row})),"")',
            )
            for col, fmt in (("L", "0.0%"), ("M", "0.0%"), ("N", "0.0%")):
                ws[f"{col}{row}"].number_format = fmt

        # Overall statistics (columns Q-T)
        ws["Q2"] = f"=AVERAGE(L2:L{last})"
        ws["R2"] = f"=_xlfn.STDEV.S(L2:L{last})"
        ws["S2"] = f"=AVERAGE(N2:N{last})"
        ws["T2"] = f"=AVERAGE(O2:O{last})"
        for coord, fmt in (("Q2", "0.0%"), ("R2", "0.0%"), ("S2", "0.0%"), ("T2", "0.00")):
            ws[coord].number_format = fmt

        try:
            wb.save(self.file_path)
        except PermissionError:
            raise MeasurementLogError(self._locked_file_message())
        except OSError as e:
            raise MeasurementLogError(f"Failed to create log file:\n{str(e)}")

    @staticmethod
    def _write_row_formulas(ws, row: int):
        """Write the throughput/loss %/loss dB formulas for one data row"""
        ws[f"E{row}"] = f'=IFERROR(D{row}/C{row},"")'
        ws[f"F{row}"] = f'=IFERROR(1-E{row},"")'
        ws[f"G{row}"] = f'=IFERROR(10*LOG10(E{row}),"")'
        ws[f"E{row}"].number_format = "0.0%"
        ws[f"F{row}"].number_format = "0.0%"
        ws[f"G{row}"].number_format = "0.00"

    def _append_xlsx(
        self,
        port: int,
        injection_power_w: Optional[float],
        output_power_w: Optional[float],
    ) -> int:
        if not self.file_path.exists():
            raise MeasurementLogError(f"Log file not found: {self.file_path}")
        try:
            wb = openpyxl.load_workbook(self.file_path)
        except PermissionError:
            raise MeasurementLogError(self._locked_file_message())
        except Exception as e:
            raise MeasurementLogError(f"Failed to open log file:\n{str(e)}")

        ws = wb.worksheets[0]

        # Find the last row containing data in columns A-D and the highest
        # trial number so far (tolerates blank spacer rows in the template)
        last_data_row = FIRST_DATA_ROW - 1
        last_trial = 0
        for row in ws.iter_rows(
            min_row=FIRST_DATA_ROW, min_col=COL_TRIAL, max_col=COL_OUTPUT
        ):
            if any(cell.value is not None for cell in row):
                last_data_row = row[0].row
            trial_value = row[0].value
            if isinstance(trial_value, (int, float)):
                last_trial = max(last_trial, int(trial_value))

        next_row = last_data_row + 1
        trial = last_trial + 1

        ws.cell(row=next_row, column=COL_TRIAL, value=trial)
        ws.cell(row=next_row, column=COL_PORT, value=port)
        if injection_power_w is not None:
            ws.cell(row=next_row, column=COL_INJECTION, value=injection_power_w)
        if output_power_w is not None:
            ws.cell(row=next_row, column=COL_OUTPUT, value=output_power_w)

        # Extend throughput/loss formulas if beyond the prefilled rows
        if ws.cell(row=next_row, column=COL_THROUGHPUT).value is None:
            self._write_row_formulas(ws, next_row)

        try:
            wb.save(self.file_path)
        except PermissionError:
            raise MeasurementLogError(self._locked_file_message())
        except OSError as e:
            raise MeasurementLogError(f"Failed to save log file:\n{str(e)}")

        logger.info(
            f"Logged trial {trial} (port {port}) to {self.file_path.name}"
        )
        return trial

    def _read_xlsx(self):
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        except PermissionError:
            raise MeasurementLogError(self._locked_file_message())
        except Exception as e:
            raise MeasurementLogError(f"Failed to open log file:\n{str(e)}")

        try:
            ws = wb.worksheets[0]
            rows = []
            for row in ws.iter_rows(
                min_row=FIRST_DATA_ROW,
                min_col=COL_TRIAL,
                max_col=COL_OUTPUT,
                values_only=True,
            ):
                if all(value is None for value in row):
                    continue
                rows.append(tuple(row))
            return rows
        finally:
            wb.close()

    # === Lantern Test Report implementation ===

    @staticmethod
    def _report_sheet(wb, wavelength_nm: Optional[int]):
        """Return the report sheet whose wavelength is nearest wavelength_nm"""
        best = None
        best_diff = None
        for ws in wb.worksheets:
            digits = "".join(ch for ch in ws.title if ch.isdigit())
            if not digits:
                continue
            if wavelength_nm is None:
                return ws
            diff = abs(int(digits) - wavelength_nm)
            if best is None or diff < best_diff:
                best, best_diff = ws, diff
        return best if best is not None else wb.worksheets[0]

    def _create_report_xlsx(self, serial: Optional[str]):
        wb = openpyxl.Workbook()
        # openpyxl writes no cached formula results; make Excel recalculate
        wb.calculation.fullCalcOnLoad = True
        wb.remove(wb.active)
        for wavelength_nm in REPORT_WAVELENGTHS_NM:
            ws = wb.create_sheet(f"{wavelength_nm} nm")
            self._write_report_sheet(ws, wavelength_nm, serial)

        try:
            wb.save(self.file_path)
        except PermissionError:
            raise MeasurementLogError(self._locked_file_message())
        except OSError as e:
            raise MeasurementLogError(f"Failed to create log file:\n{str(e)}")

    def _write_report_sheet(self, ws, wavelength_nm: int, serial: Optional[str]):
        """Write one wavelength sheet of the Lantern Test Report layout"""
        bold = Font(bold=True)
        source = "Isolated AFW Arduino Controlled Laser Diode"

        ws["A1"] = "Lantern S/N"
        ws["A1"].font = bold
        if serial:
            ws["B1"] = serial
        ws["A3"] = "Sources: "
        for row, label in ((4, "1550nm"), (5, "1310nm"), (6, "1064nm")):
            ws[f"A{row}"] = label
            ws[f"B{row}"] = source
        ws["A10"] = "Power Meters:"
        ws["A11"] = "2x PM100 USB"

        ws["A16"] = "Wavelength"
        ws["B16"] = wavelength_nm
        ws["C16"] = "nm"
        ws["A17"] = "PMREF"
        ws["C17"] = "uW"
        ws[REPORT_PMREF_CELL].number_format = "0.00"
        ws["A18"] = "Launch"
        ws["C18"] = "mW"
        ws["D18"] = "(Ref patch cord into PM1)"

        headers = {
            "A": "Port Number",
            "B": "Connector ID",
            "C": "Throughput (mW)",
            "D": "Ref (uW)",
            "E": "Calibrated Insertion Loss (dB)",
            "F": "Calibrated % Throughput",
        }
        for col, text in headers.items():
            cell = ws[f"{col}{REPORT_HEADER_ROW}"]
            cell.value = text
            cell.font = bold

        for port in range(1, REPORT_MAX_PORTS + 1):
            row = REPORT_FIRST_PORT_ROW + port - 1
            ws.cell(row=row, column=REPORT_COL_PORT, value=port)
            ws[f"E{row}"] = f"=10*LOG10(C{row}/($B$18*(D{row}/$B$17)))"
            ws[f"F{row}"] = f"=C{row}/($B$18*D{row}/$B$17)"
            ws[f"E{row}"].number_format = "0.00"
            ws[f"F{row}"].number_format = "0.00%"

        last = REPORT_FIRST_PORT_ROW + REPORT_MAX_PORTS - 1
        for offset, (label, fn) in enumerate(
            (("min", "MIN"), ("max", "MAX"), ("avg", "AVERAGE"))
        ):
            row = last + 5 + offset
            ws[f"E{row}"] = label
            ws[f"F{row}"] = f"={fn}(F{REPORT_FIRST_PORT_ROW}:F{last})"
            ws[f"F{row}"].number_format = "0.00%"

        for column, width in REPORT_COLUMN_WIDTHS.items():
            ws.column_dimensions[column].width = width

    def _append_report_xlsx(
        self,
        port: int,
        output_power_w: Optional[float],
        raw_reference_w: Optional[float],
        wavelength_nm: Optional[int],
    ) -> int:
        if not self.file_path.exists():
            raise MeasurementLogError(f"Log file not found: {self.file_path}")
        if not 1 <= port <= REPORT_MAX_PORTS:
            raise MeasurementLogError(
                f"Lantern Test Report logs support ports 1-{REPORT_MAX_PORTS} "
                f"(got {port})"
            )

        wb = self._load_workbook()
        ws = self._report_sheet(wb, wavelength_nm)
        row = REPORT_FIRST_PORT_ROW + port - 1

        ws.cell(row=row, column=REPORT_COL_PORT, value=port)
        if output_power_w is not None:
            ws.cell(row=row, column=REPORT_COL_OUTPUT_MW, value=output_power_w * 1e3)
        if raw_reference_w is not None:
            ws.cell(row=row, column=REPORT_COL_REF_UW, value=raw_reference_w * 1e6)

        try:
            wb.save(self.file_path)
        except PermissionError:
            raise MeasurementLogError(self._locked_file_message())
        except OSError as e:
            raise MeasurementLogError(f"Failed to save log file:\n{str(e)}")

        logger.info(
            f"Logged port {port} to sheet '{ws.title}' of {self.file_path.name}"
        )
        return port

    def read_report_export(self, wavelength_nm: Optional[int] = None) -> dict:
        """Collect one report sheet's data for export (e.g. to Airtable).

        Returns a dict with the report filename, lantern serial (B1),
        the sheet's wavelength (B16), pmref_uw (B17), launch_mw (B18) and
        ports: a list of (port, throughput_mW, reference_uW) tuples for
        ports that have both values recorded. Values are in the sheet's
        native units. The sheet is chosen by wavelength_nm as elsewhere.
        """
        if self.format != "xlsx" or self._resolve_layout() != "report":
            raise MeasurementLogError(
                "Export is only available for Lantern Test Report logs"
            )
        if not self.file_path.exists():
            raise MeasurementLogError(f"Log file not found: {self.file_path}")

        wb = self._load_workbook(read_only=True)
        try:
            ws = self._report_sheet(wb, wavelength_nm)
            serial = ws["B1"].value
            sheet_wavelength = ws["B16"].value
            pmref_uw = ws[REPORT_PMREF_CELL].value
            launch_mw = ws[REPORT_LAUNCH_CELL].value

            ports = []
            for port in range(1, REPORT_MAX_PORTS + 1):
                row = REPORT_FIRST_PORT_ROW + port - 1
                thru_mw = ws.cell(row=row, column=REPORT_COL_OUTPUT_MW).value
                ref_uw = ws.cell(row=row, column=REPORT_COL_REF_UW).value
                if isinstance(thru_mw, (int, float)) and isinstance(
                    ref_uw, (int, float)
                ):
                    ports.append((port, float(thru_mw), float(ref_uw)))

            return {
                "filename": self.file_path.name,
                "serial": str(serial).strip() if serial is not None else None,
                "wavelength_nm": (
                    float(sheet_wavelength)
                    if isinstance(sheet_wavelength, (int, float))
                    else None
                ),
                "pmref_uw": (
                    float(pmref_uw) if isinstance(pmref_uw, (int, float)) else None
                ),
                "launch_mw": (
                    float(launch_mw) if isinstance(launch_mw, (int, float)) else None
                ),
                "ports": ports,
            }
        finally:
            wb.close()

    def has_port_measurement(
        self, port: int, wavelength_nm: Optional[int] = None
    ) -> bool:
        """Return True if a report log already holds data for the port's row.

        Only report logs are port-keyed (logging a port again overwrites its
        row); throughput/CSV logs always append, so this returns False for
        them. The sheet is chosen by wavelength_nm as in append_measurement.
        """
        if (
            self.format != "xlsx"
            or not self.file_path.exists()
            or self._resolve_layout() != "report"
            or not 1 <= port <= REPORT_MAX_PORTS
        ):
            return False
        wb = self._load_workbook(read_only=True)
        try:
            ws = self._report_sheet(wb, wavelength_nm)
            row = REPORT_FIRST_PORT_ROW + port - 1
            return (
                ws.cell(row=row, column=REPORT_COL_OUTPUT_MW).value is not None
                or ws.cell(row=row, column=REPORT_COL_REF_UW).value is not None
            )
        finally:
            wb.close()

    def set_report_calibration(
        self, wavelength_nm: Optional[int], pmref_w: float, launch_w: float
    ):
        """Write the PMREF (uW) and Launch (mW) cells of a report sheet.

        PMREF is the raw reference reading and Launch the target reading
        taken with the reference patch cord in the target meter (i.e. the
        readings of a Calibrate Now step).
        """
        if self.format != "xlsx" or self._resolve_layout() != "report":
            raise MeasurementLogError(
                "Calibration cells are only available in Lantern Test Report logs"
            )

        wb = self._load_workbook()
        ws = self._report_sheet(wb, wavelength_nm)
        ws[REPORT_PMREF_CELL] = pmref_w * 1e6
        ws[REPORT_LAUNCH_CELL] = launch_w * 1e3

        try:
            wb.save(self.file_path)
        except PermissionError:
            raise MeasurementLogError(self._locked_file_message())
        except OSError as e:
            raise MeasurementLogError(f"Failed to save log file:\n{str(e)}")

        logger.info(
            f"Recorded PMREF/Launch calibration on sheet '{ws.title}' of "
            f"{self.file_path.name}"
        )

    def get_report_calibration(self, wavelength_nm: Optional[int] = None):
        """Return (pmref_w, launch_w) from a report sheet, or None if not set"""
        if (
            self.format != "xlsx"
            or not self.file_path.exists()
            or self._resolve_layout() != "report"
        ):
            return None
        wb = self._load_workbook(read_only=True)
        try:
            ws = self._report_sheet(wb, wavelength_nm)
            pmref_uw = ws[REPORT_PMREF_CELL].value
            launch_mw = ws[REPORT_LAUNCH_CELL].value
            if isinstance(pmref_uw, (int, float)) and isinstance(launch_mw, (int, float)):
                return pmref_uw * 1e-6, launch_mw * 1e-3
            return None
        finally:
            wb.close()

    def _read_report_xlsx(self, wavelength_nm: Optional[int]):
        wb = self._load_workbook(read_only=True)
        try:
            ws = self._report_sheet(wb, wavelength_nm)
            pmref_uw = ws[REPORT_PMREF_CELL].value
            launch_mw = ws[REPORT_LAUNCH_CELL].value
            have_calibration = (
                isinstance(pmref_uw, (int, float))
                and isinstance(launch_mw, (int, float))
                and pmref_uw > 0
            )

            rows = []
            for port in range(1, REPORT_MAX_PORTS + 1):
                row = REPORT_FIRST_PORT_ROW + port - 1
                output_mw = ws.cell(row=row, column=REPORT_COL_OUTPUT_MW).value
                ref_uw = ws.cell(row=row, column=REPORT_COL_REF_UW).value
                if output_mw is None and ref_uw is None:
                    continue

                output_w = (
                    output_mw * 1e-3 if isinstance(output_mw, (int, float)) else None
                )
                raw_ref_w = (
                    ref_uw * 1e-6 if isinstance(ref_uw, (int, float)) else None
                )
                injection_w = None
                if have_calibration and raw_ref_w is not None:
                    injection_w = (launch_mw * 1e-3) * (ref_uw / pmref_uw)
                # 5th element: the raw reference actually stored in the sheet
                rows.append((port, port, injection_w, output_w, raw_ref_w))
            return rows
        finally:
            wb.close()

    # === CSV implementation ===

    def _create_csv(self):
        try:
            with open(self.file_path, "w", newline="") as f:
                csv.writer(f).writerow(CSV_HEADER)
        except (PermissionError, OSError) as e:
            raise MeasurementLogError(f"Failed to create log file:\n{str(e)}")

    def _append_csv(
        self,
        port: int,
        injection_power_w: Optional[float],
        output_power_w: Optional[float],
        wavelength_nm: Optional[int],
    ) -> int:
        # Determine the next trial number from existing rows
        last_trial = 0
        needs_header = True
        if self.file_path.exists():
            try:
                with open(self.file_path, "r", newline="") as f:
                    for row in csv.reader(f):
                        if not row:
                            continue
                        needs_header = False
                        try:
                            last_trial = max(last_trial, int(row[1]))
                        except (IndexError, ValueError):
                            pass  # header or malformed row
            except OSError as e:
                raise MeasurementLogError(f"Failed to read log file:\n{str(e)}")

        trial = last_trial + 1

        throughput = loss_percent = loss_db = None
        if (
            injection_power_w is not None
            and output_power_w is not None
            and injection_power_w > 0
        ):
            throughput = output_power_w / injection_power_w
            loss_percent = (1.0 - throughput) * 100.0
            if throughput > 0:
                loss_db = 10.0 * math.log10(throughput)

        def fmt(value, spec="{:.6e}"):
            return spec.format(value) if value is not None else ""

        try:
            with open(self.file_path, "a", newline="") as f:
                writer = csv.writer(f)
                if needs_header:
                    writer.writerow(CSV_HEADER)
                writer.writerow(
                    [
                        datetime.now().isoformat(timespec="seconds"),
                        trial,
                        port,
                        fmt(injection_power_w),
                        fmt(output_power_w),
                        fmt(throughput, "{:.6f}"),
                        fmt(loss_percent, "{:.4f}"),
                        fmt(loss_db, "{:.4f}"),
                        wavelength_nm if wavelength_nm is not None else "",
                    ]
                )
        except (PermissionError, OSError) as e:
            raise MeasurementLogError(f"Failed to write log file:\n{str(e)}")

        logger.info(
            f"Logged trial {trial} (port {port}) to {self.file_path.name}"
        )
        return trial

    def _read_csv(self):
        rows = []
        try:
            with open(self.file_path, "r", newline="") as f:
                for row in csv.reader(f):
                    if len(row) < 5:
                        continue
                    try:
                        trial = int(row[1])
                        port = int(row[2])
                    except ValueError:
                        continue  # header or malformed row
                    injection = float(row[3]) if row[3] else None
                    output = float(row[4]) if row[4] else None
                    rows.append((trial, port, injection, output))
        except OSError as e:
            raise MeasurementLogError(f"Failed to read log file:\n{str(e)}")
        return rows

    def _locked_file_message(self) -> str:
        return (
            f"Cannot write to '{self.file_path.name}'.\n"
            "The file is probably open in Excel — close it and try again."
        )
